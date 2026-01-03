import json
from openpyxl import load_workbook
from pathlib import Path

# Configuration
XLSX_FILE = 'SG_SchoolRoll_2023.xlsx'  # Update this with your XLSX file path
OUTPUT_DIR = 'output'  # Directory where JSON files will be saved
INDENT = 2  # JSON indentation for readability (set to None for compact JSON)

def convert_cell_value(cell):
    """Convert cell value to JSON-compatible type"""
    if cell.value is None:
        return None
    # Handle dates and times
    if hasattr(cell, 'is_date') and cell.is_date:
        return cell.value.isoformat()
    return cell.value

def xlsx_to_json(xlsx_file, output_dir='output', indent=2):
    """
    Convert XLSX file to JSON files (one per sheet)
    
    Args:
        xlsx_file: Path to XLSX file
        output_dir: Directory to save JSON files
        indent: JSON indentation level (None for compact)
    """
    # Create output directory if it doesn't exist
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    # Load the workbook
    print(f"Loading workbook: {xlsx_file}")
    workbook = load_workbook(filename=xlsx_file, data_only=True)
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        print(f"\nProcessing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Get all rows
        rows = list(sheet.iter_rows(values_only=False))
        
        if not rows:
            print(f"  Sheet '{sheet_name}' is empty, skipping...")
            continue
        
        # Get headers from first row
        headers = [cell.value for cell in rows[0]]
        
        # Convert data rows to list of dictionaries
        data = []
        for row in rows[1:]:
            row_data = {}
            for header, cell in zip(headers, row):
                if header:  # Only include columns with headers
                    row_data[header] = convert_cell_value(cell)
            # Only add row if it has any non-null values
            if any(v is not None for v in row_data.values()):
                data.append(row_data)
        
        # Save to JSON file
        output_file = output_path / f"{sheet_name}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        
        print(f"  Saved {len(data)} rows to: {output_file}")
    
    workbook.close()
    print(f"\nConversion complete! JSON files saved in '{output_dir}/' directory")

def xlsx_to_json_combined(xlsx_file, output_file='combined_output.json', indent=2):
    """
    Convert all sheets in XLSX file to a single JSON file
    
    Args:
        xlsx_file: Path to XLSX file
        output_file: Output JSON file path
        indent: JSON indentation level (None for compact)
    """
    print(f"Loading workbook: {xlsx_file}")
    workbook = load_workbook(filename=xlsx_file, data_only=True)
    
    combined_data = {}
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Get all rows
        rows = list(sheet.iter_rows(values_only=False))
        
        if not rows:
            combined_data[sheet_name] = []
            continue
        
        # Get headers from first row
        headers = [cell.value for cell in rows[0]]
        
        # Convert data rows to list of dictionaries
        data = []
        for row in rows[1:]:
            row_data = {}
            for header, cell in zip(headers, row):
                if header:
                    row_data[header] = convert_cell_value(cell)
            if any(v is not None for v in row_data.values()):
                data.append(row_data)
        
        combined_data[sheet_name] = data
        print(f"  Processed {len(data)} rows from sheet '{sheet_name}'")
    
    # Save combined data to JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(combined_data, f, indent=indent, ensure_ascii=False)
    
    workbook.close()
    print(f"\nConversion complete! All sheets saved to: {output_file}")

if __name__ == '__main__':
    # Option 1: Create separate JSON files for each sheet
    xlsx_to_json(XLSX_FILE, OUTPUT_DIR, INDENT)
    
    # Option 2: Create a single JSON file with all sheets (uncomment to use)
    # xlsx_to_json_combined(XLSX_FILE, 'all_sheets.json', INDENT)



# {
#     "name": "Example University",
#     "city": "Dundee",
#     "country": "Scotland",
#     "lat": 56.462,
#     "lng": -2.9707,
#     "contacts": ["Brown"],
#     "courses": "BSc Computing Science, MSc Data Science."
#   }